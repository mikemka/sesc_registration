from core.subj_names import get_subjects
from django.http import FileResponse
from tempfile import NamedTemporaryFile
import openpyxl as xls
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from sesc.settings import SITE_URL
from string import ascii_uppercase as ALPH


GRAY_BACKGROUND = PatternFill(bgColor='E7E7E7', fill_type='solid')


def summ(iterable):
    summed = iterable[0]
    for i in iterable[1:]:
        summed += i
    return summed


def cell_center_ptserif(iterable, wrap_text=True):
    """
    Set cells' alignment to center & change font to PT Serif
    ```py
    cell_center_ptserif([sheet.cell(1, 1), sheet['A2']])
    ```
    """
    for cell in iterable:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap_text)
        cell.font = Font(name='PT Serif', size=12)
    return iterable


def cell_borders(iterable, style='thin'):
    """
    Add borders to cells
    ```py
    cell_borders([sheet.cell(1, 1), sheet['A2']])
    ```
    """
    border = Side(border_style=style, color="000000")
    for cell in iterable:
        cell.border = Border(top=border, left=border, right=border, bottom=border)
    return iterable


def col_width(iterable, width=16):
    for col in iterable:
        col.width = width
    return iterable


def cell_background(iterable, color='EDEDED'):
    for cell in iterable:
        cell.fill = PatternFill(start_color=color, fill_type='solid')
    return iterable


def export(request):
    BOOK = xls.Workbook()
    BOOK.remove_sheet(BOOK.active)
    COLORS = ('E1EFDA', 'FFF2CD', 'FFBF00', 'FFAAAA')
    form_index_gl = 0
    for f_name, full_data in (gl_subj_info := get_subjects().items()):
        form_index_gl += 1
        data, records = full_data
        sh = BOOK.create_sheet(f_name)
        # Merging
        sh.merge_cells('A2:A3')
        sh.merge_cells('B1:B3')
        sh.merge_cells(start_row=1, end_row=1, start_column=3, end_column=2 + (data_len := len(data[0]) + len(data[1])))
        sh.merge_cells(start_row=2, end_row=2, start_column=3, end_column=2 + len(data[0]))
        sh.merge_cells(start_row=2, end_row=2, start_column=3 + len(data[0]), end_column=2 + data_len)
        # Headers
        cell_center_ptserif((
            sh.cell(1, 1, f'Профиль {f_name}'),
            sh.cell(1, 2, 'класс'),
            cell_background((sh.cell(1, 3, 'Блок 2: часть, формируемая участниками образовательных отношений'),))[-1],
            sh.cell(2, 1, 'Ф.И.О. обучающегося'),
            sh.cell(2, 3, 'Учебные предметы по выбору'),
            sh.cell(2, 3 + len(data[0]), 'Элективные курсы')
        ))
        # Add subjects
        ind = 2
        for x, i in enumerate(data):
            cell_background((sh.cell(2, ind + 1),), color=COLORS[x])
            for j in i:
                cell_center_ptserif(cell_background((sh.cell(3, (ind := ind + 1), j),), color=COLORS[x]))
        # Add pupils
        stats = {}
        for ind, pupil in enumerate(records):
            cell_background((sh.cell(4 + ind, 2),))
            x, checked = 0, False
            for y, section in enumerate(data):
                for subject in section:
                    stats.setdefault(subject, 0)
                    if section[subject] == 1 or pupil in section[subject]:
                        if section[subject] != 1:
                            checked = True
                        stats[subject] += 1
                        sh.cell(4 + ind, 3 + x, 1)
                    cell_background((sh.cell(4 + ind, 3 + x),), color=COLORS[y])
                    x += 1
            cell_center_ptserif(cell_borders([sh.cell(4 + ind, i) for i in range(1, 3 + data_len)]))
            cell_background((sh.cell(4 + ind, 1, pupil.name),), color=COLORS[1 if checked else 3]),
            sh.cell(4 + ind, 1).hyperlink = f'{SITE_URL}/{form_index_gl}/{pupil.slug}'
        # Add statistics
        sh.merge_cells(start_row=4 + len(records), end_row=4 + len(records), start_column=1, end_column=2)
        sh.cell(4 + len(records), 1, 'всего человек в группе')
        for i, j in enumerate(stats):
            sh.cell(4 + len(records), 3 + i, stats[j])
        # Other styles
        cell_borders(summ([[sh.cell(i, 3 + j) for j in range(data_len)] for i in range(2, 4)]))
        cell_center_ptserif(cell_borders(cell_background([sh.cell(4 + len(records), i) for i in range(1, data_len + 3)], color=COLORS[2])))
        col_width([sh.column_dimensions[ALPH[(2 + i) % len(ALPH)]] for i in range(data_len)])
        col_width((sh.column_dimensions['A'],), width=24.33)
    # Global statistics
    stats = [{} for _ in range(2)]
    for f_name, info in gl_subj_info:
        info, all_pupils = info
        for pos, block in enumerate(info):
            for subject, s_info in block.items():
                stats[pos][subject] = stats[pos].setdefault(subject, 0) + s_info.count() if s_info != 1 else all_pupils.count()
    sh, x = BOOK.create_sheet('свод по группам'), 2
    for color, block in enumerate(stats):
        for subject, num in block.items():
            cell_borders(cell_center_ptserif([sh.cell(x, 2, num), sh.cell(x, 3, num / 10)] + cell_background([sh.cell(x, 1, subject)], color=COLORS[color])))
            sh.row_dimensions[x].height = 50
            x += 1
    # Styles
    cell_center_ptserif(cell_background((sh.cell(1, 1), sh.cell(1, 2), sh.cell(1, 3, 'количество групп (10 человек в группе)'))))
    col_width((sh.column_dimensions['A'],), width=24)
    col_width((sh.column_dimensions['B'],), width=10)
    col_width((sh.column_dimensions['C'],))
    cell_borders((sh.cell(1, 3),))
    # Sending .xlsx
    book_file = NamedTemporaryFile(prefix='exp_', suffix='.xlsx')
    BOOK.save(book_file.name)
    export_response = FileResponse(open(book_file.name, 'rb'), as_attachment=True)
    book_file.close()
    return export_response
